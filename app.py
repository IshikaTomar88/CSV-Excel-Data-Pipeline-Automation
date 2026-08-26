"""
UNIVERSAL ENTERPRISE DATA PIPELINE & EXECUTIVE REPORTING SUITE
Supports: Digital Marketing Agencies, Real Estate Brokers, Shopify & Skin Clinics
"""

import io
from dataclasses import dataclass, field
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# --------------------------------------------------------------------------
# PAGE CONFIGURATION & STYLING
# --------------------------------------------------------------------------
st.set_page_config(
    page_title="Universal Enterprise Pipeline & Executive Reporter",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
    <style>
        .main { background-color: #0f172a; color: #f8fafc; }
        .stMetric { background-color: #1e293b; padding: 15px; border-radius: 10px; border: 1px solid #334155; }
        .stTabs [data-baseweb="tab-list"] { gap: 8px; }
        .stTabs [data-baseweb="tab"] { background-color: #1e293b; border-radius: 6px; color: #cbd5e1; padding: 10px 20px; font-weight: 600; }
        .stTabs [aria-selected="true"] { background-color: #2563eb !important; color: white !important; }
        div.stButton > button { background-color: #2563eb; color: white; font-weight: bold; border-radius: 6px; padding: 0.5rem 1rem; border: none; width: 100%; }
        div.stButton > button:hover { background-color: #1d4ed8; }
    </style>
    """,
    unsafe_allow_html=True,
)

@dataclass
class DataQualityReport:
    rows_in: int = 0
    rows_out: int = 0
    duplicates_removed: int = 0
    missing_values_filled: dict[str, int] = field(default_factory=dict)
    type_coercion_failures: dict[str, int] = field(default_factory=dict)
    outliers_flagged: int = 0

    def summary_lines(self) -> list[str]:
        lines = [
            f"Total rows ingested: {self.rows_in}",
            f"Total rows in final clean dataset: {self.rows_out}",
            f"Exact duplicate rows dropped: {self.duplicates_removed}",
            f"Statistical outliers flagged (IQR method): {self.outliers_flagged}",
        ]
        for col, n in self.missing_values_filled.items():
            lines.append(f"  - Column '{col}': {n} missing values filled safely")
        for col, n in self.type_coercion_failures.items():
            lines.append(f"  - Column '{col}': {n} malformed values coerced to NaN")
        return lines


class UniversalEnterprisePipeline:
    def __init__(self, primary_file, secondary_file=None, mapping=None):
        self.primary_file = primary_file
        self.secondary_file = secondary_file
        self.mapping = mapping or {}
        self.report = DataQualityReport()
        self.df_clean: pd.DataFrame | None = None

    def _load_file(self, file_obj) -> pd.DataFrame:
        filename = file_obj.name.lower()
        if filename.endswith(".csv"):
            return pd.read_csv(file_obj)
        elif filename.endswith((".xls", ".xlsx")):
            return pd.read_excel(file_obj)
        else:
            raise ValueError(f"Unsupported file format for {file_obj.name}")

    def process(self) -> pd.DataFrame:
        df = self._load_file(self.primary_file)
        self.report.rows_in = len(df)

        # Optional multi-source merge
        if self.secondary_file is not None:
            df2 = self._load_file(self.secondary_file)
            self.report.rows_in += len(df2)
            common_cols = list(set(df.columns).intersection(set(df2.columns)))
            if common_cols:
                join_key = common_cols[0]
                df = pd.merge(df, df2, on=join_key, how="outer", suffixes=("_primary", "_secondary"))
            else:
                df = pd.concat([df, df2], ignore_index=True, sort=False)

        # Bulletproof Explicit Column Mapping
        rev_source = self.mapping.get("revenue")
        if rev_source and rev_source in df.columns and rev_source != "None":
            df["revenue"] = df[rev_source]
        else:
            for col in df.columns:
                if any(x in col.lower() for x in ["rev", "sales", "amount", "total", "price", "value"]):
                    df["revenue"] = df[col]
                    break

        if "revenue" not in df.columns:
            raise ValueError(f"Could not map a valid Revenue/Sales column. Available columns: {list(df.columns)}")

        # Handle optional mapped fields
        for target_key, map_key in [
            ("order_date", "order_date"), 
            ("spend", "spend"), 
            ("leads", "leads"), 
            ("region", "region")
        ]:
            source_col = self.mapping.get(map_key)
            if source_col and source_col in df.columns and source_col != "None":
                if source_col != target_key:
                    df[target_key] = df[source_col]

        # Clean string columns
        for col in df.select_dtypes(include=["object", "string"]).columns:
            df[col] = df[col].astype(str).str.strip().replace({"nan": np.nan, "": np.nan, "None": np.nan})

        # AGGRESSIVE NUMERIC COERCION (Fixes the $0.00 bug by cleaning symbols & parsing correctly)
        for col in ["revenue", "spend", "leads", "quantity", "unit_price"]:
            if col in df.columns:
                before_na = df[col].isna().sum()
                # Remove currency symbols, commas, and whitespace completely
                cleaned = df[col].astype(str).str.replace(r"[$,€£₹USD\s]", "", regex=True).str.replace(",", "", regex=False)
                df[col] = pd.to_numeric(cleaned, errors="coerce")
                failures = df[col].isna().sum() - before_na
                if failures > 0:
                    self.report.type_coercion_failures[col] = int(failures)

        # Fallback safeguard: If revenue became all NaNs/zeros due to mapping error, try finding the first numeric column
        if df["revenue"].sum() == 0:
            for c in df.select_dtypes(include=[np.number]).columns:
                if c != "revenue":
                    df["revenue"] = df[c]
                    break

        # Parse dates safely
        if "order_date" in df.columns:
            df["order_date"] = pd.to_datetime(df["order_date"], errors="coerce")

        # Compute performance metrics
        if "spend" in df.columns and "revenue" in df.columns:
            df["spend"] = df["spend"].fillna(0)
            df["roas"] = np.where(df["spend"] > 0, df["revenue"] / df["spend"], 0)
        if "spend" in df.columns and "leads" in df.columns:
            df["leads"] = df["leads"].fillna(0)
            df["true_cpa"] = np.where(df["leads"] > 0, df["spend"] / df["leads"], 0)

        # Impute missing cells safely
        for col in df.columns:
            n_missing = int(df[col].isna().sum())
            if n_missing == 0:
                continue
            if pd.api.types.is_numeric_dtype(df[col]):
                df[col] = df[col].fillna(0.0) # Default to 0 for missing numbers
                self.report.missing_values_filled[col] = n_missing
            elif col != "order_date":
                df[col] = df[col].fillna("UNKNOWN")
                self.report.missing_values_filled[col] = n_missing

        # Strip exact duplicates
        before = len(df)
        df = df.drop_duplicates()
        self.report.duplicates_removed = before - len(df)

        # Flag outliers using IQR method
        if "revenue" in df.columns and len(df) > 5:
            q1, q3 = df["revenue"].quantile([0.25, 0.75])
            iqr = q3 - q1
            lower, upper = q1 - 1.5 * iqr, q3 + 1.5 * iqr
            df["flag_outlier"] = ~df["revenue"].between(lower, upper)
            self.report.outliers_flagged = int(df["flag_outlier"].sum())
        else:
            df["flag_outlier"] = False

        self.df_clean = df.reset_index(drop=True)
        self.report.rows_out = len(self.df_clean)
        return self.df_clean

    def analyze(self) -> dict:
        df = self.df_clean
        summary = {}
        if "revenue" in df.columns:
            summary["total_revenue"] = round(float(df["revenue"].sum()), 2)
            summary["avg_order_value"] = round(float(df["revenue"].mean()), 2)
            summary["order_count"] = int(len(df))
        if "spend" in df.columns:
            summary["total_spend"] = round(float(df["spend"].sum()), 2)
            summary["overall_roas"] = round(summary["total_revenue"] / summary["total_spend"], 2) if summary["total_spend"] > 0 else 0.0
        if "leads" in df.columns:
            summary["total_leads"] = int(df["leads"].sum())
            summary["blended_cpa"] = round(summary.get("total_spend", 0) / summary["total_leads"], 2) if summary["total_leads"] > 0 else 0.0

        group_candidates = [c for c in df.select_dtypes(include=["object", "string"]).columns if c != "flag_outlier"]
        group_col = group_candidates[0] if group_candidates else None
        
        if group_col and "revenue" in df.columns:
            by_group = df.groupby(group_col)["revenue"].sum().sort_values(ascending=False).round(2)
            total = by_group.sum()
            summary["group_column_name"] = group_col
            summary["revenue_by_group"] = by_group
            summary["revenue_by_group_pct"] = (by_group / total * 100).round(1) if total else by_group * 0

        if "order_date" in df.columns and df["order_date"].notna().sum() > 0:
            monthly = df.dropna(subset=["order_date"]).set_index("order_date")["revenue"].resample("ME").sum().round(2)
            summary["monthly_revenue"] = monthly

        return summary

    def build_excel_bytes(self, summary: dict) -> bytes:
        wb = Workbook()
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2F5496")
        subheader_font = Font(bold=True)
        currency_fmt = '"$"#,##0.00'
        pct_fmt = "0.0%"

        ws = wb.active
        ws.title = "Executive Summary"
        ws["A1"] = "Executive Business & Performance Summary"
        ws["A1"].font = header_font
        ws["A1"].fill = header_fill
        ws.merge_cells("A1:D1")

        row = 3
        ws[f"A{row}"] = "Metric"
        ws[f"B{row}"] = "Value"
        for c in ("A", "B"): ws[f"{c}{row}"].font = subheader_font
        row += 1

        metrics = [
            ("total_revenue", "Total Revenue / Sales", currency_fmt),
            ("total_spend", "Total Ad Spend", currency_fmt),
            ("overall_roas", "Blended ROAS", "ROAS_FMT"),
            ("total_leads", "Total Leads / Volume", None),
            ("blended_cpa", "Blended CPA", currency_fmt),
            ("avg_order_value", "Average Order Value", currency_fmt),
            ("order_count", "Total Clean Records", None)
        ]
        for key, label, fmt in metrics:
            if key in summary:
                ws[f"A{row}"] = label
                cell = ws[f"B{row}"]
                cell.value = summary[key]
                if fmt and fmt != "ROAS_FMT": 
                    cell.number_format = fmt
                elif fmt == "ROAS_FMT": 
                    cell.number_format = '0.00"x"'
                row += 1

        group_col = summary.get("group_column_name", "Group").title()
        if "revenue_by_group" in summary:
            row += 2
            ws[f"A{row}"] = f"Revenue Breakdown by {group_col}"
            ws[f"A{row}"].font = subheader_font
            row += 1
            ws[f"A{row}"], ws[f"B{row}"], ws[f"C{row}"] = group_col, "Revenue", "% of Total"
            for c in ("A", "B", "C"): ws[f"{c}{row}"].font = subheader_font
            row += 1
            chart_start = row
            pct = summary.get("revenue_by_group_pct")
            for name, val in summary["revenue_by_group"].items():
                ws[f"A{row}"] = str(name)
                ws[f"B{row}"] = val
                ws[f"B{row}"].number_format = currency_fmt
                if pct is not None and name in pct:
                    ws[f"C{row}"] = float(pct[name]) / 100
                    ws[f"C{row}"].number_format = pct_fmt
                row += 1
            chart_end = row - 1

            chart = BarChart()
            chart.title = f"Revenue by {group_col}"
            chart.y_axis.title = "Revenue ($)"
            chart.add_data(Reference(ws, min_col=2, min_row=chart_start, max_row=chart_end), titles_from_data=False)
            chart.set_categories(Reference(ws, min_col=1, min_row=chart_start, max_row=chart_end))
            ws.add_chart(chart, f"E{chart_start}")
            row += 2

        if "monthly_revenue" in summary and not summary["monthly_revenue"].empty:
            ws[f"A{row}"] = "Monthly Revenue Trend"
            ws[f"A{row}"].font = subheader_font
            row += 1
            ws[f"A{row}"], ws[f"B{row}"] = "Month", "Revenue"
            for c in ("A", "B"): ws[f"{c}{row}"].font = subheader_font
            row += 1
            trend_start = row
            for month, val in summary["monthly_revenue"].items():
                ws[f"A{row}"] = month.strftime("%Y-%m")
                ws[f"B{row}"] = val
                ws[f"B{row}"].number_format = currency_fmt
                row += 1
            trend_end = row - 1

            line = LineChart()
            line.title = "Monthly Revenue Trend"
            line.add_data(Reference(ws, min_col=2, min_row=trend_start, max_row=trend_end), titles_from_data=False)
            line.set_categories(Reference(ws, min_col=1, min_row=trend_start, max_row=trend_end))
            ws.add_chart(line, f"E{trend_start}")

        for col, width in zip("ABCDE", (28, 16, 14, 4, 45)):
            ws.column_dimensions[col].width = width

        ws2 = wb.create_sheet("Data Quality Report")
        ws2["A1"] = "Pipeline Data Quality & Health Audit"
        ws2["A1"].font = header_font
        ws2["A1"].fill = header_fill
        ws2.merge_cells("A1:B1")
        for i, line in enumerate(self.report.summary_lines(), start=3):
            ws2[f"A{i}"] = line
        ws2.column_dimensions["A"].width = 75

        ws3 = wb.create_sheet("Cleaned Raw Data")
        for r in dataframe_to_rows(self.df_clean, index=False, header=True):
            ws3.append(r)
        for cell in ws3[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F5496")
        for i, _ in enumerate(self.df_clean.columns, start=1):
            ws3.column_dimensions[get_column_letter(i)].width = 16
        ws3.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


# STREAMLIT USER INTERFACE
st.sidebar.title("⚡ Enterprise Studio")
st.sidebar.markdown("---")
st.sidebar.info("Universal data janitor and executive report builder for agencies, real estate brokers, and e-commerce stores.")
st.sidebar.markdown("### Developer")
st.sidebar.write("**Ishika Tomar**")
st.sidebar.markdown("[🔗 Connect on LinkedIn](https://www.linkedin.com/in/ishika-tomar-70262a2a5/)")

st.title("📊 Universal Enterprise Business & Marketing Pipeline")
st.markdown("Transform messy exports from Shopify, Facebook Ads, Real Estate CRMs, or Skin Clinic booking sheets into polished, executive-ready Excel reports instantly.")

col_up1, col_up2 = st.columns(2)
with col_up1:
    primary_file = st.file_uploader("Upload Primary File (Sales / Revenue CSV or Excel)", type=["csv", "xlsx", "xls"])
with col_up2:
    secondary_file = st.file_uploader("Upload Secondary File [Optional] (Ad Spend / CRM Leads)", type=["csv", "xlsx", "xls"])

if primary_file is not None:
    try:
        temp_df = pd.read_csv(primary_file) if primary_file.name.endswith(".csv") else pd.read_excel(primary_file)
        if secondary_file is not None:
            temp_df2 = pd.read_csv(secondary_file) if secondary_file.name.endswith(".csv") else pd.read_excel(secondary_file)
            cols = list(set(temp_df.columns).union(set(temp_df2.columns)))
        else:
            cols = list(temp_df.columns)

        st.markdown("### 🗺️ Interactive Column Mapper")
        st.info("Map your file columns to standard reporting fields so the pipeline processes seamlessly:")
        
        def_rev_idx = 0
        for idx, c in enumerate(cols):
            if any(x in c.lower() for x in ["rev", "sales", "amount", "total", "price", "value"]):
                def_rev_idx = idx
                break

        m1, m2, m3 = st.columns(3)
        with m1:
            rev_col = st.selectbox("Revenue / Sales Column (Required)", cols, index=def_rev_idx)
            date_col = st.selectbox("Date Column (Optional)", ["None"] + cols)
        with m2:
            spend_col = st.selectbox("Ad Spend Column [Optional]", ["None"] + cols)
            leads_col = st.selectbox("Leads / Quantity Column [Optional]", ["None"] + cols)
        with m3:
            region_col = st.selectbox("Region / Group / Category Column [Optional]", ["None"] + cols)

        mapping = {
            "revenue": rev_col,
            "order_date": None if date_col == "None" else date_col,
            "spend": None if spend_col == "None" else spend_col,
            "leads": None if leads_col == "None" else leads_col,
            "region": None if region_col == "None" else region_col
        }

        st.markdown("---")
        if st.button("🚀 Execute Enterprise Pipeline & Generate Report"):
            with st.spinner("Processing data, cleaning anomalies, and formatting executive workbook..."):
                pipeline = UniversalEnterprisePipeline(primary_file, secondary_file, mapping)
                df_clean = pipeline.process()
                summary = pipeline.analyze()
                excel_bytes = pipeline.build_excel_bytes(summary)

            st.success("✅ Pipeline Executed Successfully!")

            c1, c2, c3, c4, c5 = st.columns(5)
            with c1:
                st.metric("Total Revenue", f"${summary.get('total_revenue', 0):,.2f}")
            with c2:
                st.metric("Total Ad Spend", f"${summary.get('total_spend', 0):,.2f}")
            with c3:
                st.metric("Blended ROAS", f"{summary.get('overall_roas', 0):.2f}x")
            with c4:
                st.metric("Blended CPA", f"${summary.get('blended_cpa', 0):,.2f}")
            with c5:
                st.metric("Outliers Flagged", f"{pipeline.report.outliers_flagged}" if 'pipeline' in locals() else "0")

            st.markdown("---")
            tab1, tab2, tab3 = st.tabs(["📊 Performance Breakdown", "🧹 Data Quality Audit Log", "🔍 Cleaned Dataset Preview"])

            with tab1:
                st.subheader("Executive Performance Overview")
                if "revenue_by_group" in summary:
                    col_a, col_b = st.columns(2)
                    with col_a:
                        st.write(f"**Revenue Breakdown by {summary.get('group_column_name', 'Group').title()}**")
                        st.dataframe(summary["revenue_by_group"], use_container_width=True)
                    with col_b:
                        if "monthly_revenue" in summary and not summary["monthly_revenue"].empty:
                            st.write("**Monthly Revenue Trend**")
                            st.line_chart(summary["monthly_revenue"])
                else:
                    st.info("Clean dataset processed successfully. Map a grouping/region column to view category distribution charts.")

            with tab2:
                st.subheader("Data Health & Integrity Audit Log")
                for line in pipeline.report.summary_lines():
                    st.text(line)

            with tab3:
                st.subheader("Cleaned Dataset Preview")
                st.dataframe(df_clean, use_container_width=True)

            st.markdown("---")
            st.subheader("📥 Export Client Deliverable")
            st.download_button(
                label="Download Formatted Executive Report (.xlsx)",
                data=excel_bytes,
                file_name="universal_enterprise_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    except Exception as e:
        st.error(f"Pipeline Error: {str(e)}")
else:
    st.markdown(
        """
        > **Get Started:** Upload your primary data file (Shopify export, Real Estate sheet, or Agency ad report). Optionally upload a secondary file to merge them instantly!
        """
    )
